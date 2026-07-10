"""往来单位管理路由。"""

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional

from app.database import get_db
from app.models import Counterparty
from app.schemas import CounterpartyResponse
from app.auth import get_current_user

router = APIRouter()


class CounterpartyUpdate(BaseModel):
    name: Optional[str] = None
    short_name: Optional[str] = None
    category: Optional[str] = None
    tax_number: Optional[str] = None
    bank_account: Optional[str] = None
    bank_name: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    contact_person: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    zip_code: Optional[str] = None


@router.get("/", response_model=list[CounterpartyResponse])
def list_counterparties(company_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    return (
        db.query(Counterparty)
        .filter(Counterparty.company_id == company_id, Counterparty.is_active)
        .order_by(Counterparty.code)
        .all()
    )


@router.post("/", response_model=CounterpartyResponse)
def create_counterparty(
    company_id: int, data: CounterpartyUpdate, db: Session = Depends(get_db), user=Depends(get_current_user)
):
    import random

    code = str(random.randint(10000, 99999))
    while db.query(Counterparty).filter(Counterparty.company_id == company_id, Counterparty.code == code).first():
        code = str(random.randint(10000, 99999))
    cp = Counterparty(company_id=company_id, code=code, name=data.name or "")
    for field in [
        "short_name",
        "category",
        "tax_number",
        "bank_account",
        "bank_name",
        "address",
        "phone",
        "contact_person",
        "website",
        "email",
        "zip_code",
    ]:
        val = getattr(data, field, None)
        if val is not None:
            setattr(cp, field, val)
    db.add(cp)
    db.commit()
    db.refresh(cp)
    return cp


@router.put("/{counterparty_id}", response_model=CounterpartyResponse)
def update_counterparty(
    counterparty_id: int, data: CounterpartyUpdate, db: Session = Depends(get_db), user=Depends(get_current_user)
):
    cp = db.query(Counterparty).filter(Counterparty.id == counterparty_id).first()
    if not cp:
        raise HTTPException(status_code=404, detail="客户不存在")
    for field in [
        "name",
        "short_name",
        "category",
        "tax_number",
        "bank_account",
        "bank_name",
        "address",
        "phone",
        "contact_person",
        "website",
        "email",
        "zip_code",
    ]:
        val = getattr(data, field, None)
        if val is not None:
            setattr(cp, field, val)
    db.commit()
    db.refresh(cp)
    return cp


@router.delete("/{counterparty_id}")
def delete_counterparty(counterparty_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    cp = db.query(Counterparty).filter(Counterparty.id == counterparty_id).first()
    if not cp:
        raise HTTPException(status_code=404, detail="客户不存在")
    cp.is_active = False
    db.commit()
    return {"ok": True}


@router.post("/import")
def import_counterparties(company_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """从 basic_master_data/往来单位.xlsx 导入。"""
    from pathlib import Path

    try:
        import openpyxl
    except ImportError as err:
        raise HTTPException(status_code=500, detail="需要安装 openpyxl") from err

    data_dir = Path(__file__).parent.parent.parent.parent / "basic_master_data"
    wb_path = data_dir / "往来单位.xlsx"
    if not wb_path.exists():
        raise HTTPException(status_code=404, detail=f"找不到文件: {wb_path}")

    wb = openpyxl.load_workbook(wb_path)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        code = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ""
        if not code or not name:
            continue
        existing = (
            db.query(Counterparty).filter(Counterparty.company_id == company_id, Counterparty.code == code).first()
        )
        if not existing:
            db.add(
                Counterparty(
                    company_id=company_id,
                    code=code,
                    name=name,
                    short_name=str(row[2]).strip() if row[2] else None,
                    category=str(row[3]).strip() if row[3] else None,
                    category_code=str(row[4]).strip() if row[4] else None,
                )
            )
            count += 1
    db.commit()
    wb.close()
    return {"ok": True, "imported": count}
