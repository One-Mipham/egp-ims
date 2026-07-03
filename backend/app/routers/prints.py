"""打印模块路由：公司信息、部门、科目表、科目余额表、总账、凭证、月/季/年报。"""
import calendar
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import User, Company, Department, Account, Voucher, VoucherEntry, AccountingPeriod, Counterparty, Person, Project
from app.auth import get_current_user
from app.routers.reports import (
    _period_end_date, _year_start, _prev_year_period, _prev_year_end,
    _calc_ending, _occurrence,
    BS_ROWS, IS_ROWS,
    _compute_cash_flows, _is_cash_account,
    CASH_CODES,
)

router = APIRouter()


def _get_company(db: Session, company_id: int) -> Company:
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="公司不存在")
    return company


# ──────────────── 公司信息 ────────────────

@router.get("/company")
def print_company(company_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    company = _get_company(db, company_id)
    return {
        "name": company.name,
        "short_name": company.short_name,
        "industry": company.industry,
        "currency": company.currency,
        "fiscal_year_start": company.fiscal_year_start,
        "internal_control_mode": company.internal_control_mode,
    }


# ──────────────── 部门信息 ────────────────

@router.get("/departments")
def print_departments(company_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    _get_company(db, company_id)
    depts = db.query(Department).filter(
        Department.company_id == company_id,
    ).order_by(Department.code).all()
    return [
        {"code": d.code, "name": d.name, "manager": d.manager, "is_active": d.is_active}
        for d in depts
    ]


# ──────────────── 科目表 ────────────────

@router.get("/subjects")
def print_subjects(
    company_id: int,
    level: Optional[int] = Query(None, ge=1, le=4),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _get_company(db, company_id)
    q = db.query(Account).filter(Account.company_id == company_id, Account.is_active == True)
    if level is not None:
        q = q.filter(Account.level == level)
    accts = q.order_by(Account.code).all()
    LEVEL_LABELS = {1: "一级科目", 2: "二级科目", 3: "三级科目", 4: "四级科目"}
    return [
        {
            "code": a.code,
            "name": a.name,
            "level": LEVEL_LABELS.get(a.level, str(a.level)),
            "category": a.category,
            "balance_direction": a.balance_direction,
            "parent_code": a.parent_code,
        }
        for a in accts
    ]


# ──────────────── 科目余额表 ────────────────

@router.get("/subject-balance")
def print_subject_balance(
    company_id: int,
    period: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _get_company(db, company_id)
    end_date = _period_end_date(period)
    prev_period_end = _period_end_date(f"{int(period[:4])}-{int(period[5:7]) - 1:02d}" if int(period[5:7]) > 1 else f"{int(period[:4]) - 1}-12")

    accts = db.query(Account).filter(
        Account.company_id == company_id,
        Account.is_active == True,
    ).order_by(Account.code).all()

    rows = []
    for a in accts:
        beginning = _calc_ending(a, db, company_id, period[:4] + "-01-01")
        # 本期发生
        d, c = _occurrence(a, db, company_id, period + "-01", end_date)
        ending = _calc_ending(a, db, company_id, end_date)
        rows.append({
            "code": a.code,
            "name": a.name,
            "beginning": round(beginning, 2),
            "debit": round(d, 2),
            "credit": round(c, 2),
            "ending": round(ending, 2),
        })

    return {"period": period, "rows": rows}


# ──────────────── 总账余额表 ────────────────

@router.get("/general-ledger")
def print_general_ledger(
    company_id: int,
    period: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _get_company(db, company_id)
    end_date = _period_end_date(period)
    # 只取一级科目
    accts = db.query(Account).filter(
        Account.company_id == company_id,
        Account.level == 1,
        Account.is_active == True,
    ).order_by(Account.code).all()

    rows = []
    for a in accts:
        beginning = _calc_ending(a, db, company_id, period[:4] + "-01-01")
        d, c = _occurrence(a, db, company_id, period + "-01", end_date)
        ending = _calc_ending(a, db, company_id, end_date)
        rows.append({
            "code": a.code,
            "name": a.name,
            "beginning": round(beginning, 2),
            "debit": round(d, 2),
            "credit": round(c, 2),
            "ending": round(ending, 2),
        })

    return {"period": period, "rows": rows}


# ──────────────── 凭证打印 ────────────────

@router.get("/vouchers")
def print_vouchers(
    company_id: int,
    range: str = Query("month", pattern="^(today|week|month|custom)$"),
    start_date: str = Query(None),
    end_date: str = Query(None),
    voucher_no_from: str = Query(None),
    voucher_no_to: str = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _get_company(db, company_id)
    now = datetime.now(timezone(timedelta(hours=8)))
    today = now.strftime("%Y-%m-%d")
    if range == "custom" and start_date and end_date:
        start, end = start_date, end_date
    elif range == "today":
        start, end = today, today
    elif range == "week":
        monday = now - timedelta(days=now.weekday())
        start = monday.strftime("%Y-%m-%d")
        end = today
    else:
        start = now.strftime("%Y-%m") + "-01"
        end = today

    q = db.query(Voucher).filter(
        Voucher.company_id == company_id,
        Voucher.date >= start,
        Voucher.date <= end,
    )
    if voucher_no_from:
        q = q.filter(Voucher.voucher_no >= voucher_no_from)
    if voucher_no_to:
        q = q.filter(Voucher.voucher_no <= voucher_no_to)
    vouchers = q.order_by(Voucher.date).all()

    result = []
    for v in vouchers:
        entries = db.query(VoucherEntry).filter(VoucherEntry.voucher_id == v.id).all()
        entry_data = []
        for e in entries:
            dept_name = db.query(Department).filter(Department.id == e.department_id).first() if e.department_id else None
            cp_name = db.query(Counterparty).filter(Counterparty.id == e.counterparty_id).first() if e.counterparty_id else None
            p_name = db.query(Person).filter(Person.id == e.person_id).first() if e.person_id else None
            pr_name = db.query(Project).filter(Project.id == e.project_id).first() if e.project_id else None
            entry_data.append({
                "account_code": e.account_code,
                "debit": e.debit,
                "credit": e.credit,
                "description": e.description,
                "department_name": dept_name.name if dept_name else None,
                "counterparty_name": cp_name.name if cp_name else None,
                "person_name": p_name.name if p_name else None,
                "project_name": pr_name.name if pr_name else None,
            })
        result.append({
            "id": v.id,
            "date": v.date,
            "voucher_no": v.voucher_no,
            "voucher_type": v.voucher_type,
            "summary": v.summary,
            "status": v.status,
            "entries": entry_data,
        })
    return {"range": range, "start": start, "end": end, "vouchers": result}


# ──────────────── 月报/季报/年报 ────────────────

def _period_range(period: str, rtype: str) -> tuple[str, str]:
    """Return (curr_start, prev_start) based on period type.
    - monthly: curr=当月, prev=去年同月
    - quarterly: curr=当季首月, prev=去年同季
    - yearly: curr=当年1月, prev=去年1月
    """
    y, m = int(period[:4]), int(period[5:7])
    if rtype == "quarterly":
        qm = ((m - 1) // 3) * 3 + 1  # first month of quarter
        return f"{y}-{qm:02d}-01", f"{y - 1}-{qm:02d}-01"
    elif rtype == "yearly":
        return f"{y}-01-01", f"{y - 1}-01-01"
    else:  # monthly
        return f"{period}-01", f"{_prev_year_period(period)}-01"


def _prev_day(date_str: str) -> str:
    """Return the day before a YYYY-MM-DD date string.
    e.g. '2026-05-01' → '2026-04-30' (last day of previous month).
    Used to get period-beginning balance (balance at end of prior day = start of this day)."""
    from datetime import timedelta as _td
    y, m, d = int(date_str[:4]), int(date_str[5:7]), int(date_str[8:10])
    dt = datetime(y, m, d) - _td(days=1)
    return dt.strftime("%Y-%m-%d")


def _get_report_data(db: Session, company_id: int, period: str, report: str, rtype: str = "monthly"):
    end_date = _period_end_date(period)
    ys = _year_start(period)
    py_end = _period_end_date(_prev_year_period(period))
    py_ys = _year_start(_prev_year_period(period))

    # For quarterly/yearly, the "current period" spans multiple months
    curr_start, prev_curr_start = _period_range(period, rtype)

    accts = {a.code: a for a in db.query(Account).filter(Account.company_id == company_id).all()}
    parent_codes = set()
    for a in accts.values():
        if a.parent_code:
            parent_codes.add(a.parent_code)

    if report == "balance":
        def _calc(code_str, ref_date=None):
            if not code_str or code_str in ("CURRENT_TOTAL", "NCURRENT_TOTAL", "ASSET_TOTAL", "LIABILITY_TOTAL", "EQUITY_TOTAL", "TOTAL"):
                return 0.0
            codes = [c.strip() for c in code_str.split(",") if c.strip()]
            total = 0.0
            target_date = ref_date if ref_date is not None else end_date
            for code in codes:
                children = [a for a in accts.values() if a.code.startswith(code) and a.code not in parent_codes]
                if not children and code in accts and code not in parent_codes:
                    children = [accts[code]]
                for a in children:
                    total += _calc_ending(a, db, company_id, target_date)
            return round(total, 2)

        pye = _prev_year_end(period)
        left_items, right_items = [], []
        for name, side, codes in BS_ROWS:
            val = _calc(codes)
            item = {"name": name, "ending": val, "beginning": _calc(codes, pye)}
            if side == "left":
                left_items.append(item)
            else:
                right_items.append(item)

        for item in left_items:
            if item["name"] == "流动资产合计":
                item["ending"] = round(sum(i["ending"] for i in left_items[:10]), 2)
                item["beginning"] = round(sum(i["beginning"] for i in left_items[:10]), 2)
            elif item["name"] == "非流动资产合计":
                item["ending"] = round(sum(i["ending"] for i in left_items[11:29]), 2)
                item["beginning"] = round(sum(i["beginning"] for i in left_items[11:29]), 2)
            elif item["name"] == "资产总计":
                item["ending"] = round(left_items[10]["ending"] + left_items[29]["ending"], 2)
                item["beginning"] = round(left_items[10]["beginning"] + left_items[29]["beginning"], 2)

        for item in right_items:
            if item["name"] == "流动负债合计":
                item["ending"] = round(sum(i["ending"] for i in right_items[:12]), 2)
                item["beginning"] = round(sum(i["beginning"] for i in right_items[:12]), 2)
            elif item["name"] == "非流动负债合计":
                item["ending"] = round(sum(i["ending"] for i in right_items[13:21]), 2)
                item["beginning"] = round(sum(i["beginning"] for i in right_items[13:21]), 2)
            elif item["name"] == "负债合计":
                item["ending"] = round(right_items[12]["ending"] + right_items[21]["ending"], 2)
                item["beginning"] = round(right_items[12]["beginning"] + right_items[21]["beginning"], 2)
            elif item["name"] == "所有者权益合计":
                item["ending"] = round(sum(i["ending"] for i in right_items[24:29]), 2)
                item["beginning"] = round(sum(i["beginning"] for i in right_items[24:29]), 2)
            elif item["name"] == "负债和所有者权益总计":
                item["ending"] = round(right_items[22]["ending"] + right_items[29]["ending"], 2)
                item["beginning"] = round(right_items[22]["beginning"] + right_items[29]["beginning"], 2)

        # ── 倒挤平衡：将尾差吸入未分配利润 ──
        asset_total = left_items[30]["ending"]
        liability_equity_total = right_items[30]["ending"]
        diff = round(asset_total - liability_equity_total, 2)
        if abs(diff) > 0.005:
            for item in right_items:
                if item["name"] == "未分配利润":
                    item["ending"] = round(item["ending"] + diff, 2)
                    item["beginning"] = round(item["beginning"] + diff, 2)
                    break
            for item in right_items:
                if item["name"] == "所有者权益合计":
                    item["ending"] = round(sum(i["ending"] for i in right_items[24:29]), 2)
                    item["beginning"] = round(sum(i["beginning"] for i in right_items[24:29]), 2)
                elif item["name"] == "负债和所有者权益总计":
                    item["ending"] = round(right_items[22]["ending"] + right_items[29]["ending"], 2)
                    item["beginning"] = round(right_items[22]["beginning"] + right_items[29]["beginning"], 2)

        y, m = int(period[:4]), int(period[5:7])
        if rtype == "yearly":
            date_display = f"{y} 年 12 月 31 日"
        elif rtype == "quarterly":
            q = (m - 1) // 3 + 1
            qm_last = {1: 31, 2: 30, 3: 30, 4: 31}
            date_display = f"{y} 年第{q}季度（{y} 年 {m:02d} 月 {qm_last.get(q, 31)} 日）"
        else:
            last_day = calendar.monthrange(y, m)[1]
            date_display = f"{y} 年 {m:02d} 月 {last_day} 日"
        return {"type": "balance", "date_display": date_display, "left_items": left_items, "right_items": right_items}

    elif report == "income":
        def _calc(code_str, start, end):
            if not code_str:
                return 0.0
            codes = [c.strip() for c in code_str.split(",") if c.strip()]
            total = 0.0
            for code in codes:
                children = [a for a in accts.values() if a.code.startswith(code) and a.code not in parent_codes]
                if not children and code in accts and code not in parent_codes:
                    children = [accts[code]]
                for a in children:
                    exclude_xfer = (a.category == "profit_loss")
                    d, c = _occurrence(a, db, company_id, start, end, exclude_transfer=exclude_xfer)
                    if a.category == "profit_loss":
                        if a.code.startswith(("64", "66", "67", "68")):
                            total += d
                        else:
                            total += c
            return round(total, 2)

        # curr_start 根据 rtype 覆盖单月/季度/年度起始
        items = []
        for name, codes in IS_ROWS:
            if codes in ("OP_PROFIT", "TOTAL_PROFIT", "NET_PROFIT"):
                items.append({"name": name, "curr": 0, "ytd": 0, "prev": 0, "formula": codes})
            else:
                items.append({
                    "name": name,
                    "curr": round(_calc(codes, curr_start, end_date), 2),
                    "ytd": round(_calc(codes, ys, end_date), 2),
                    "prev": round(_calc(codes, prev_curr_start, py_end), 2),
                    "formula": "",
                })

        for item in items:
            f = item.get("formula", "")
            if f == "OP_PROFIT":
                op = items[0]["curr"] - items[1]["curr"] - items[2]["curr"] - items[3]["curr"] - items[4]["curr"] - items[5]["curr"] - items[6]["curr"]
                item["curr"] = round(op, 2)
            elif f == "TOTAL_PROFIT":
                op = next(i["curr"] for i in items if i.get("formula") == "OP_PROFIT")
                nonop = items[10]["curr"] - items[11]["curr"]
                item["curr"] = round(op + nonop, 2)
            elif f == "NET_PROFIT":
                tp = next(i["curr"] for i in items if i.get("formula") == "TOTAL_PROFIT")
                item["curr"] = round(tp - items[13]["curr"], 2)

        for col in ("ytd", "prev"):
            s = ys if col == "ytd" else prev_curr_start
            e = end_date if col == "ytd" else py_end
            def _calc_col(code_str):
                if not code_str:
                    return 0.0
                return _calc(code_str, s, e)
            vals = {}
            for item in items:
                if item.get("formula", "") in ("OP_PROFIT", "TOTAL_PROFIT", "NET_PROFIT"):
                    continue
                vals[item["name"]] = round(_calc_col(next((codes for n, codes in IS_ROWS if n == item["name"]), "")), 2)
            for item in items:
                f = item.get("formula", "")
                if f == "OP_PROFIT":
                    item[col] = round(vals["一、营业收入"] - vals["    减：营业成本"] - vals["        税金及附加"] - vals["        销售费用"] - vals["        管理费用"] - vals["        财务费用"] - vals["        资产减值损失"], 2)
                elif f == "TOTAL_PROFIT":
                    op = next(i[col] for i in items if i.get("formula") == "OP_PROFIT")
                    item[col] = round(op + vals["    加：营业外收入"] - vals["    减：营业外支出"], 2)
                elif f == "NET_PROFIT":
                    tp = next(i[col] for i in items if i.get("formula") == "TOTAL_PROFIT")
                    item[col] = round(tp - vals["    减：所得税费用"], 2)

        y, m = int(period[:4]), int(period[5:7])
        if rtype == "yearly":
            period_display = f"{y} 年度"
        elif rtype == "quarterly":
            q = (m - 1) // 3 + 1
            period_display = f"{y} 年第{q}季度"
        else:
            period_display = f"{y} 年 {m:02d} 月"
        return {"type": "income", "period_display": period_display, "items": items}

    elif report == "cashflow":
        from app.models import CashFlowItem

        cf_items = {cfi.code: cfi for cfi in db.query(CashFlowItem).filter(
            CashFlowItem.company_id == company_id, CashFlowItem.is_active == True
        ).all()}

        curr_items = _compute_cash_flows(db, company_id, curr_start, end_date)  # 本期(月度/季度/年度)
        ytd_items = _compute_cash_flows(db, company_id, ys, end_date)  # 本年累计
        prev_items = _compute_cash_flows(db, company_id, prev_curr_start, py_end)  # 上年同期

        def _cat_sum(data: dict, prefix: str, direction: str) -> float:
            return sum(v for k, v in data.items() if cf_items.get(k) and cf_items[k].category_code
                       and cf_items[k].category_code.startswith(prefix) and cf_items[k].direction == direction)

        # Curr period
        op_in_c = round(_cat_sum(curr_items, "op_", "inflow"), 2)
        op_out_c = round(_cat_sum(curr_items, "op_", "outflow"), 2)
        inv_in_c = round(_cat_sum(curr_items, "inv_", "inflow"), 2)
        inv_out_c = round(_cat_sum(curr_items, "inv_", "outflow"), 2)
        fin_in_c = round(_cat_sum(curr_items, "fin_", "inflow"), 2)
        fin_out_c = round(_cat_sum(curr_items, "fin_", "outflow"), 2)
        # YTD
        op_in_y = round(_cat_sum(ytd_items, "op_", "inflow"), 2)
        op_out_y = round(_cat_sum(ytd_items, "op_", "outflow"), 2)
        inv_in_y = round(_cat_sum(ytd_items, "inv_", "inflow"), 2)
        inv_out_y = round(_cat_sum(ytd_items, "inv_", "outflow"), 2)
        fin_in_y = round(_cat_sum(ytd_items, "fin_", "inflow"), 2)
        fin_out_y = round(_cat_sum(ytd_items, "fin_", "outflow"), 2)
        # Prev
        op_in_p = round(_cat_sum(prev_items, "op_", "inflow"), 2)
        op_out_p = round(_cat_sum(prev_items, "op_", "outflow"), 2)
        inv_in_p = round(_cat_sum(prev_items, "inv_", "inflow"), 2)
        inv_out_p = round(_cat_sum(prev_items, "inv_", "outflow"), 2)
        fin_in_p = round(_cat_sum(prev_items, "fin_", "inflow"), 2)
        fin_out_p = round(_cat_sum(prev_items, "fin_", "outflow"), 2)

        # 现金类科目（与资产负债表货币资金一致：1001,1002,1012）
        all_accts = db.query(Account).filter(Account.company_id == company_id).all()
        accts_cf = {a.code: a for a in all_accts if _is_cash_account(a.code, CASH_CODES)}
        pye = _prev_year_end(period)

        def _cash_balance(ref_date):
            """资产负债表货币资金在某一日期的余额"""
            return round(sum(_calc_ending(a, db, company_id, ref_date) for a in accts_cf.values()), 2)

        # ── 倒挤法：期末余额 = 资产负债表货币资金 ──
        # curr 列：本期期初 = curr_start 前一日；期末 = end_date
        beg_c = _cash_balance(_prev_day(curr_start))
        end_c = _cash_balance(end_date)
        # ytd 列：期初 = 去年12月31日；期末 = end_date（同curr）
        beg_y = _cash_balance(pye)
        end_y = end_c  # ytd 和 curr 期末是同一天（本期期末）
        # prev 列：上年同期期初 = prev_curr_start 前一日；期末 = py_end
        beg_p = _cash_balance(_prev_day(prev_curr_start))
        end_p = _cash_balance(py_end)

        def _net(inc, outc): return round(inc - outc, 2)
        net_op_c = _net(op_in_c, op_out_c); net_op_y = _net(op_in_y, op_out_y); net_op_p = _net(op_in_p, op_out_p)
        net_inv_c = _net(inv_in_c, inv_out_c); net_inv_y = _net(inv_in_y, inv_out_y); net_inv_p = _net(inv_in_p, inv_out_p)
        net_fin_c = _net(fin_in_c, fin_out_c); net_fin_y = _net(fin_in_y, fin_out_y); net_fin_p = _net(fin_in_p, fin_out_p)
        # 净变动 = 期末 - 期初（倒挤，保证与资产负债表对齐）
        net_chg_c = round(end_c - beg_c, 2)
        net_chg_y = round(end_y - beg_y, 2)
        net_chg_p = round(end_p - beg_p, 2)

        y, m = int(period[:4]), int(period[5:7])
        if rtype == "yearly":
            date_display = f"{y} 年度"
        elif rtype == "quarterly":
            q = (m - 1) // 3 + 1
            date_display = f"{y} 年第{q}季度"
        else:
            last_day = calendar.monthrange(y, m)[1]
            date_display = f"{y} 年 {m:02d} 月 {last_day} 日"

        rows = [
            ("一、经营活动产生的现金流量：", "", "", ""),
            ("     销售商品、提供劳务收到的现金", op_in_c, op_in_y, op_in_p),
            ("     经营活动现金流入小计", op_in_c, op_in_y, op_in_p),
            ("     支付其它与经营活动有关的现金", op_out_c, op_out_y, op_out_p),
            ("     经营活动现金流出小计", op_out_c, op_out_y, op_out_p),
            ("     经营活动产生的现金流量净额", net_op_c, net_op_y, net_op_p),
            ("二、投资活动产生的现金流量：", "", "", ""),
            ("     收回投资收到的现金", inv_in_c, inv_in_y, inv_in_p),
            ("     投资活动现金流入小计", inv_in_c, inv_in_y, inv_in_p),
            ("     投资支付的现金", inv_out_c, inv_out_y, inv_out_p),
            ("     投资活动现金流出小计", inv_out_c, inv_out_y, inv_out_p),
            ("     投资活动产生的现金流量净额", net_inv_c, net_inv_y, net_inv_p),
            ("三、筹资活动产生的现金流量：", "", "", ""),
            ("     吸收投资收到的现金", fin_in_c, fin_in_y, fin_in_p),
            ("     筹资活动现金流入小计", fin_in_c, fin_in_y, fin_in_p),
            ("     支付其他与筹资活动有关的现金", fin_out_c, fin_out_y, fin_out_p),
            ("     筹资活动现金流出小计", fin_out_c, fin_out_y, fin_out_p),
            ("     筹资活动产生的现金流量净额", net_fin_c, net_fin_y, net_fin_p),
            ("五、现金及现金等价物净增加额", net_chg_c, net_chg_y, net_chg_p),
            ("        加：期初现金及现金等价物余额", beg_c, beg_y, beg_p),
            ("六、期末现金及现金等价物余额", end_c, end_y, end_p),
        ]
        return {"type": "cashflow", "date_display": date_display, "rows": rows}

    raise HTTPException(status_code=400, detail="报表类型不支持")


@router.get("/periodic")
def print_periodic(
    company_id: int,
    period: str,
    report: str = Query("balance", pattern="^(balance|income|cashflow)$"),
    type: str = Query("monthly", pattern="^(monthly|quarterly|yearly)$"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _get_company(db, company_id)
    return _get_report_data(db, company_id, period, report, type)


@router.get("/periodic/export")
def export_periodic(
    company_id: int,
    period: str,
    report: str = Query("balance", pattern="^(balance|income|cashflow)$"),
    type: str = Query("monthly", pattern="^(monthly|quarterly|yearly)$"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导出月/季/年度报表为 Excel 文件。"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from fastapi.responses import StreamingResponse

    _get_company(db, company_id)
    data = _get_report_data(db, company_id, period, report, type)

    wb = Workbook()
    ws = wb.active
    ws.title = {"balance": "资产负债表", "income": "利润表", "cashflow": "现金流量表"}[report]

    # 样式
    title_font = Font(name="微软雅黑", size=16, bold=True)
    header_font = Font(name="微软雅黑", size=10, bold=True)
    normal_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="F5F5F4", end_color="F5F5F4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    number_align = Alignment(horizontal="right", vertical="center")
    text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    TYPE_LABEL = {"monthly": "月度", "quarterly": "季度", "yearly": "年度"}
    type_label = TYPE_LABEL.get(type, "")

    row = 1

    if report == "balance":
        # 标题
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="资产负债表").font = title_font
        ws.cell(row=row, column=1).alignment = center_align
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"{type_label}报表 — {data['date_display']}").font = normal_font
        ws.cell(row=row, column=1).alignment = center_align
        row += 2

        # 表头
        headers = ["资产", "期末余额", "年初余额", "负债及所有者权益", "期末余额", "年初余额"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = header_font; c.alignment = center_align; c.border = thin_border; c.fill = header_fill
        row += 1

        # 数据行
        for i in range(max(len(data["left_items"]), len(data["right_items"]))):
            for ci in range(1, 7):
                ws.cell(row=row, column=ci).border = thin_border
                ws.cell(row=row, column=ci).font = normal_font
            left = data["left_items"][i] if i < len(data["left_items"]) else None
            right = data["right_items"][i] if i < len(data["right_items"]) else None
            is_bold = (left and ("合计" in left["name"] or "总计" in left["name"])) or \
                      (right and ("合计" in right["name"] or "总计" in right["name"]))

            if left:
                ws.cell(row=row, column=1, value=left["name"]).alignment = text_align
                ws.cell(row=row, column=2, value=left["ending"]).alignment = number_align
                ws.cell(row=row, column=3, value=left["beginning"]).alignment = number_align
            if right:
                ws.cell(row=row, column=4, value=right["name"]).alignment = text_align
                ws.cell(row=row, column=5, value=right["ending"]).alignment = number_align
                ws.cell(row=row, column=6, value=right["beginning"]).alignment = number_align
            if is_bold:
                for ci in range(1, 7):
                    ws.cell(row=row, column=ci).font = header_font
                    ws.cell(row=row, column=ci).fill = header_fill
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16

    elif report == "income":
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value="利润表").font = title_font
        ws.cell(row=row, column=1).alignment = center_align
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f"{type_label}报表 — {data['period_display']}").font = normal_font
        ws.cell(row=row, column=1).alignment = center_align
        row += 2

        headers = ["项目", "本期金额", "本年累计", "上年同期"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = header_font; c.alignment = center_align; c.border = thin_border; c.fill = header_fill
        row += 1

        for item in data["items"]:
            is_bold = any(k in item["name"] for k in ["营业利润", "利润总额", "净利润"])
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).border = thin_border
                ws.cell(row=row, column=ci).font = header_font if is_bold else normal_font
                if is_bold:
                    ws.cell(row=row, column=ci).fill = header_fill
            ws.cell(row=row, column=1, value=item["name"]).alignment = text_align
            ws.cell(row=row, column=2, value=item.get("curr", 0)).alignment = number_align
            ws.cell(row=row, column=3, value=item.get("ytd", 0)).alignment = number_align
            ws.cell(row=row, column=4, value=item.get("prev", 0)).alignment = number_align
            row += 1

        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

    elif report == "cashflow":
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value="现金流量表").font = title_font
        ws.cell(row=row, column=1).alignment = center_align
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f"{type_label}报表 — {data['date_display']}").font = normal_font
        ws.cell(row=row, column=1).alignment = center_align
        row += 2

        headers = ["项目", "本期金额", "本年累计", "上年同期"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = header_font; c.alignment = center_align; c.border = thin_border; c.fill = header_fill
        row += 1

        for r in data["rows"]:
            is_header = r[0].startswith(("一、", "二、", "三、", "四、", "五、", "六、"))
            is_sub = "小计" in r[0] or "净额" in r[0]
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).border = thin_border
                ws.cell(row=row, column=ci).font = header_font if (is_header or is_sub) else normal_font
                if is_header or is_sub:
                    ws.cell(row=row, column=ci).fill = header_fill
            ws.cell(row=row, column=1, value=r[0]).alignment = text_align
            for ci in range(1, 4):
                val = r[ci]
                ws.cell(row=row, column=ci + 1, value=val if isinstance(val, (int, float)) else 0).alignment = number_align
            row += 1

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

    # 输出
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{type_label}报表_{report}_{period}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
