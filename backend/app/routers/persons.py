"""员工个人管理路由。"""
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import User, Person
from app.schemas import PersonResponse
from app.auth import get_current_user

router = APIRouter()


@router.get("/", response_model=list[PersonResponse])
def list_persons(company_id: int, department_code: str = Query(None), db: Session = Depends(get_db), user=Depends(get_current_user)):
    q = db.query(Person).filter(Person.company_id == company_id, Person.is_active == True)
    if department_code:
        q = q.filter(Person.department_code == department_code)
    return q.order_by(Person.code).all()


@router.post("/import")
def import_persons(company_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """从 basic_master_data/员工.xlsx 导入。"""
    from pathlib import Path
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="需要安装 openpyxl")

    data_dir = Path(__file__).parent.parent.parent.parent / "basic_master_data"
    wb_path = data_dir / "员工.xlsx"
    if not wb_path.exists():
        raise HTTPException(status_code=404, detail=f"找不到文件: {wb_path}")

    wb = openpyxl.load_workbook(wb_path)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        code = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ""
        if not code or not name:
            continue
        dept_code = str(row[4]).strip() if row[4] else None
        existing = db.query(Person).filter(Person.company_id == company_id, Person.code == code).first()
        if not existing:
            db.add(Person(company_id=company_id, code=code, name=name, department_code=dept_code))
            count += 1
    db.commit()
    wb.close()
    return {"ok": True, "imported": count}
