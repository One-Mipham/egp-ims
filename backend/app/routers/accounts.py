"""科目管理路由。"""
from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Account, VoucherEntry, Voucher, Company
from app.schemas import AccountCreate, AccountResponse
from app.auth import get_current_user
from app.permissions import check_account_level

router = APIRouter()


class InitialBalanceUpdate(BaseModel):
    initial_balance: float


class BulkInitialBalanceItem(BaseModel):
    code: str
    initial_balance: float


class BulkInitialBalanceRequest(BaseModel):
    company_id: int
    accounts: list[BulkInitialBalanceItem]


@router.get("/", response_model=list[AccountResponse])
def list_accounts(company_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    return db.query(Account).filter(Account.company_id == company_id).order_by(Account.code).all()


@router.post("/import-aux-config")
def import_aux_config(company_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """从 basic_master_data/科目.xlsx 导入辅助核算配置。"""
    from pathlib import Path
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="需要安装 openpyxl")

    data_dir = Path(__file__).parent.parent.parent.parent / "basic_master_data"
    wb_path = data_dir / "科目.xlsx"
    if not wb_path.exists():
        raise HTTPException(status_code=404, detail=f"找不到文件: {wb_path}")

    wb = openpyxl.load_workbook(wb_path)
    ws = wb.active
    updated = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[2]:
            continue
        code = str(row[2]).strip()
        if not code:
            continue
        acct = db.query(Account).filter(Account.company_id == company_id, Account.code == code).first()
        if acct:
            acct.aux_dept = int(row[24] or 0)
            acct.aux_person = int(row[25] or 0)
            acct.aux_counterparty = int(row[26] or 0)
            updated += 1
    db.commit()
    wb.close()
    return {"ok": True, "updated": updated}


@router.get("/balance")
def account_balance(company_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """科目余额表：期初余额 + 本期借贷发生额 = 期末余额。"""
    accounts = db.query(Account).filter(Account.company_id == company_id).order_by(Account.code).all()
    result = []
    for acct in accounts:
        # 计算已记账凭证的借贷方发生额
        entries = db.query(VoucherEntry).join(Voucher).filter(
            Voucher.company_id == company_id,
            VoucherEntry.account_code == acct.code,
            Voucher.status == "posted",
        ).all()
        debit_total = sum(e.debit for e in entries)
        credit_total = sum(e.credit for e in entries)

        if acct.balance_direction == "debit":
            ending_balance = acct.initial_balance + debit_total - credit_total
        else:
            ending_balance = acct.initial_balance + credit_total - debit_total

        result.append({
            "code": acct.code,
            "name": acct.name,
            "level": acct.level,
            "category": acct.category,
            "balance_direction": acct.balance_direction,
            "initial_balance": acct.initial_balance,
            "debit_total": debit_total,
            "credit_total": credit_total,
            "ending_balance": round(ending_balance, 2),
        })
    return result


@router.post("/", response_model=AccountResponse)
def create_account(data: AccountCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    company = db.query(Company).filter(Company.id == data.company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="公司不存在")
    err = check_account_level(user, company, data.level)
    if err:
        raise HTTPException(status_code=403, detail=err)
    if data.parent_code:
        parent = db.query(Account).filter(
            Account.company_id == data.company_id,
            Account.code == data.parent_code,
        ).first()
        if not parent:
            raise HTTPException(status_code=400, detail="父科目不存在")
    account = Account(
        company_id=data.company_id, code=data.code, name=data.name, level=data.level,
        parent_code=data.parent_code, category=data.category, balance_direction=data.balance_direction,
        initial_balance=data.initial_balance,
    )
    db.add(account)
    db.commit()
    db.refresh(account)
    return account


@router.put("/{account_id}", response_model=AccountResponse)
def update_account(account_id: int, name: str = None, is_active: bool = None, db: Session = Depends(get_db), user=Depends(get_current_user)):
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="科目不存在")
    if account.is_system:
        raise HTTPException(status_code=403, detail="系统科目不可修改")
    company = db.query(Company).filter(Company.id == account.company_id).first()
    err = check_account_level(user, company, account.level)
    if err:
        raise HTTPException(status_code=403, detail=err)
    if name:
        account.name = name
    if is_active is not None:
        account.is_active = is_active
    db.commit()
    db.refresh(account)
    return account


@router.patch("/{account_id}/initial-balance")
def set_initial_balance(account_id: int, data: InitialBalanceUpdate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """设置科目期初余额。"""
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="科目不存在")
    account.initial_balance = data.initial_balance
    db.commit()
    return {"ok": True, "code": account.code, "initial_balance": account.initial_balance}


@router.post("/bulk-initial-balance")
def bulk_set_initial_balance(data: BulkInitialBalanceRequest, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """批量设置科目期初余额。"""
    updated = 0
    for item in data.accounts:
        account = db.query(Account).filter(
            Account.company_id == data.company_id,
            Account.code == item.code,
        ).first()
        if account:
            account.initial_balance = item.initial_balance
            updated += 1
    db.commit()
    return {"ok": True, "updated": updated}


@router.delete("/{account_id}")
def delete_account(account_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """删除科目。系统科目不可删，有子科目的不可删，已被凭证引用的不可删。"""
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="科目不存在")
    if account.is_system:
        raise HTTPException(status_code=403, detail="系统预置科目不可删除")

    company = db.query(Company).filter(Company.id == account.company_id).first()
    err = check_account_level(user, company, account.level)
    if err:
        raise HTTPException(status_code=403, detail=err)

    child = db.query(Account).filter(
        Account.company_id == account.company_id,
        Account.parent_code == account.code,
    ).first()
    if child:
        raise HTTPException(status_code=400, detail="该科目下存在子科目，请先删除子科目")

    used = db.query(VoucherEntry).filter(VoucherEntry.account_code == account.code).first()
    if used:
        raise HTTPException(status_code=400, detail="该科目已被凭证引用，不可删除")

    db.delete(account)
    db.commit()
    return {"ok": True}


# ── 银行结算方式 ──

SETTLEMENT_METHODS = [
    {"code": "997", "name": "转账"},
    {"code": "998", "name": "微信"},
    {"code": "999", "name": "支付宝"},
    {"code": "1", "name": "现金"},
    {"code": "2", "name": "支票"},
    {"code": "3", "name": "银行汇票"},
    {"code": "4", "name": "银行本票"},
    {"code": "5", "name": "信用卡"},
    {"code": "6", "name": "电汇"},
    {"code": "7", "name": "网银"},
    {"code": "8", "name": "自动扣费"},
    {"code": "9", "name": "委托收款"},
    {"code": "10", "name": "托收承付"},
    {"code": "11", "name": "信用证"},
    {"code": "12", "name": "其他"},
]


@router.get("/settlement-methods")
def list_settlement_methods(user=Depends(get_current_user)):
    """返回银行结算方式列表（基准数据 + 国标补充）。"""
    return SETTLEMENT_METHODS
