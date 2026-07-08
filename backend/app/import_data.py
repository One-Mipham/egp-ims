"""导入利美融信公司数据：公司、科目（4级）、期初余额、部门。"""
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from app.database import SessionLocal, init_db
from app.models import Account, Company, Department

DOWNLOADS = Path.home() / "Downloads" / "2026-05"
ACCOUNTS_XLSX = DOWNLOADS / "科目.xlsx"
BALANCES_XLSX = DOWNLOADS / "2026.01.01-2026.4.30-科目发生额及余额表.xlsx"

CATEGORY_MAP = {
    "资产": "asset",
    "负债": "liability",
    "权益": "equity",
    "成本": "cost",
    "损益": "profit_loss",
}

DIRECTION_MAP = {
    "借方": "debit",
    "贷方": "credit",
}

# 部门数据
DEPARTMENTS = [
    {"code": "001", "name": "董事会"},
    {"code": "002", "name": "顾问办公室"},
    {"code": "003", "name": "总裁办公室"},
    {"code": "004", "name": "行政部"},
    {"code": "005", "name": "人力资源部"},
    {"code": "006", "name": "财务部"},
    {"code": "008", "name": "法务事务部"},
    {"code": "009", "name": "国际业务部"},
    {"code": "010", "name": "审计稽核部"},
    {"code": "011", "name": "外汇管理部"},
    {"code": "012", "name": "并购事业部"},
    {"code": "013", "name": "咨询事业部"},
    {"code": "014", "name": "技术事业部"},
    {"code": "015", "name": "房地产业务部"},
    {"code": "016", "name": "酒店事业部"},
]


def parent_code(code: str) -> str | None:
    if len(code) <= 4:
        return None
    if len(code) == 6:
        return code[:4]
    if len(code) == 8:
        return code[:6]
    if len(code) == 10:
        return code[:8]
    return None


def create_company(db: Session) -> int:
    """创建或查找公司。"""
    name = "青岛利美融信投资控股有限责任公司"
    company = db.query(Company).filter(Company.name == name).first()
    if company:
        print(f"  公司已存在 id={company.id}")
        return company.id
    company = Company(
        name=name,
        short_name="利美融信",
        industry="investment",
        internal_control_mode="standard",
        currency="CNY",
    )
    db.add(company)
    db.commit()
    db.refresh(company)
    print(f"  创建公司: {company.name} (id={company.id})")
    return company.id


def import_accounts(db: Session, company_id: int) -> int:
    """从科目.xlsx 导入全部科目。"""
    db.query(Account).filter(Account.company_id == company_id).delete()
    db.commit()

    wb = openpyxl.load_workbook(ACCOUNTS_XLSX, read_only=True)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        level_raw, code, name, category_raw, direction_raw, _, _, disabled = row
        if not code or str(code).strip() == "":
            continue

        code = str(code).strip()
        name = str(name).strip()
        category = CATEGORY_MAP.get(str(category_raw).strip(), "asset")
        direction = DIRECTION_MAP.get(str(direction_raw).strip(), "debit")
        level = int(level_raw) if level_raw else (len(code) // 2 + 1 if len(code) > 4 else 1)

        acct = Account(
            company_id=company_id,
            code=code,
            name=name,
            level=level,
            parent_code=parent_code(code) if level > 1 else None,
            category=category,
            balance_direction=direction,
            initial_balance=0.0,
            is_system=level == 1,
            is_active=not disabled,
        )
        db.add(acct)
        count += 1

    db.commit()
    wb.close()
    print(f"  科目: {count} 条")
    return count


def import_balances(db: Session, company_id: int) -> int:
    """从 2026.01-2026.04 余额表导入期初余额（2026-01-01）。"""
    wb = openpyxl.load_workbook(BALANCES_XLSX, read_only=True)
    ws = wb.active

    # 数据行 9~83
    # 列索引: 1=类别, 2=科目编码, 3=名称, 5=期初借方, 6=期初贷方
    updated = 0
    for row in ws.iter_rows(min_row=9, max_row=83, values_only=True):
        if row is None:
            continue
        code = row[2]
        if not code or str(code).strip() == "":
            continue
        code = str(code).strip()
        beg_debit = row[5]
        beg_credit = row[6]

        beg_d = float(beg_debit) if beg_debit else 0.0
        beg_c = float(beg_credit) if beg_credit else 0.0

        if not beg_d and not beg_c:
            continue

        # 按余额方向计算期初余额
        acct = db.query(Account).filter(
            Account.company_id == company_id, Account.code == code
        ).first()
        if acct:
            # 余额表中，期初余额列显示的是该方向的发生额累计
            # 余额 = 该方向累计 - 反方向累计
            if acct.balance_direction == "debit":
                acct.initial_balance = round(beg_d - beg_c, 2)
            else:
                acct.initial_balance = round(beg_c - beg_d, 2)
            updated += 1

    db.commit()
    wb.close()
    print(f"  期初余额: {updated} 个科目")
    return updated


def import_departments(db: Session, company_id: int) -> int:
    """导入部门。"""
    db.query(Department).filter(Department.company_id == company_id).delete()
    db.commit()

    for dept in DEPARTMENTS:
        db.add(Department(
            company_id=company_id,
            code=dept["code"],
            name=dept["name"],
            is_active=True,
        ))
    db.commit()
    print(f"  部门: {len(DEPARTMENTS)} 条")
    return len(DEPARTMENTS)


def main():
    print("=" * 50)
    print("利美融信数据导入")
    print("=" * 50)
    init_db()
    db = SessionLocal()
    try:
        company_id = create_company(db)
        print()
        import_accounts(db, company_id)
        import_balances(db, company_id)
        import_departments(db, company_id)
        print()
        print("=" * 50)
        print("导入完成！")
        print("=" * 50)
    finally:
        db.close()


if __name__ == "__main__":
    main()
