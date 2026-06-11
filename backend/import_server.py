#!/usr/bin/env python3
"""Import real vouchers on the server."""
import sys, os
os.chdir("/opt/egp-ims/intranet/backend")
sys.path.insert(0, ".")

from app.database import SessionLocal
from app.models import Voucher, VoucherEntry, Account, Company
import openpyxl
from collections import defaultdict

SRC = "/opt/egp-ims/intranet/data/凭证.xlsx"
CID = 1

wb = openpyxl.load_workbook(SRC)
ws = wb.active

# Group rows into monthly sections based on voucher number restarts
sections = []
curr_rows = []
prev_num = 0

for row_idx in range(2, ws.max_row + 1):
    vno = str(ws.cell(row=row_idx, column=2).value or "").strip()
    if not vno or "合计" in vno:
        continue
    try:
        num = int(vno.split("-")[1])
    except:
        continue

    if num < prev_num and prev_num > 0:
        if curr_rows:
            sections.append(curr_rows)
        curr_rows = []
    curr_rows.append(row_idx)
    prev_num = num

if curr_rows:
    sections.append(curr_rows)

print(f"Sections found: {len(sections)}")
for i, sec in enumerate(sections):
    vnos = set()
    for r in sec:
        vnos.add(str(ws.cell(row=r, column=2).value or "").strip())
    print(f"  Section {i+1}: {len(sec)} rows, {len(vnos)} vouchers, samples: {sorted(vnos)[:3]}")

# Import
db = SessionLocal()

# Account lookup
accts = {}
for a in db.query(Account).filter(Account.company_id == CID).all():
    accts[a.code] = a

# Delete old vouchers
for v in db.query(Voucher).filter(Voucher.company_id == CID).all():
    db.query(VoucherEntry).filter(VoucherEntry.voucher_id == v.id).delete()
db.query(Voucher).filter(Voucher.company_id == CID).delete()
db.commit()
print("Old vouchers deleted")

created = 0
total_entries = 0
unknown = set()

for sec_idx, rows in enumerate(sections):
    month = sec_idx + 1
    if month > 6:
        break

    vgroups = defaultdict(list)
    for r in rows:
        vno = str(ws.cell(row=r, column=2).value or "").strip()
        vgroups[vno].append(r)

    print(f"2026-{month:02d}: {len(vgroups)} vouchers")

    day = 1
    for vno in sorted(vgroups.keys()):
        entries = []
        for r in vgroups[vno]:
            code = str(ws.cell(row=r, column=7).value or "").strip()
            debit = float(ws.cell(row=r, column=10).value or 0)
            credit = float(ws.cell(row=r, column=11).value or 0)
            summary = str(ws.cell(row=r, column=6).value or "").strip()

            if code not in accts:
                unknown.add(code)
                continue

            entries.append({"code": code, "debit": debit, "credit": credit, "summary": summary})

        if not entries:
            continue

        date = f"2026-{month:02d}-{day:02d}"
        day = min(day + 1, 28)

        v = Voucher(
            company_id=CID,
            date=date,
            voucher_no=f"记字2026{month:02d}-{vno.split('-')[1].zfill(4)}",
            voucher_type="transfer",
            summary=entries[0]["summary"],
            creator_id=1,
            status="draft",
        )
        db.add(v)
        db.flush()

        for e in entries:
            db.add(VoucherEntry(
                voucher_id=v.id,
                account_code=e["code"],
                debit=e["debit"],
                credit=e["credit"],
                description=e["summary"],
            ))
            total_entries += 1

        created += 1

db.commit()

# Verify
vouchers = db.query(Voucher).filter(Voucher.company_id == CID).all()
all_entries = []
for v in vouchers:
    all_entries.extend(db.query(VoucherEntry).filter(VoucherEntry.voucher_id == v.id).all())
td = sum(e.debit for e in all_entries)
tc = sum(e.credit for e in all_entries)

print("")
print("=" * 50)
print(f"Import done: {created} vouchers, {total_entries} entries")
print(f"Total: D={td:,.2f} C={tc:,.2f} {'OK' if abs(td-tc)<0.01 else 'IMBALANCE!'}")
if unknown:
    print(f"Unknown codes ({len(unknown)}): {sorted(unknown)[:30]}")

# Monthly summary
for m in range(1, 6):
    mvs = [v for v in vouchers if v.date.startswith(f"2026-{m:02d}")]
    mes = []
    for v in mvs:
        mes.extend(db.query(VoucherEntry).filter(VoucherEntry.voucher_id == v.id).all())
    md = sum(e.debit for e in mes)
    mc = sum(e.credit for e in mes)
    print(f"  2026-{m:02d}: {len(mvs)} vouchers D={md:>14,.2f} C={mc:>14,.2f} {'OK' if abs(md-mc)<0.01 else 'IMBAL!'}")

db.close()
