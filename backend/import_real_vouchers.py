#!/usr/bin/env python3
"""
导入 2026年1-5月 真实凭证数据。
来源: Downloads/20260101-20260531-填制凭证.xlsx
"""

import os, sys

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from collections import defaultdict

from app.database import SessionLocal
from app.models import Voucher, VoucherEntry, Company, Account

SRC = os.path.expanduser("~/Downloads/20260101-20260531-填制凭证.xlsx")
COMPANY_ID = 1


def parse_sections(ws):
    """Detect monthly sections based on voucher number restarts."""
    sections = []
    prev_num = 0
    current_rows = []

    for row_idx in range(2, ws.max_row + 1):
        vno = str(ws.cell(row=row_idx, column=2).value or "").strip()
        if not vno:
            continue
        try:
            vno_num = int(vno.split("-")[1])
        except (ValueError, IndexError):
            continue

        if vno_num < prev_num and prev_num > 0:
            sections.append(current_rows)
            current_rows = []

        current_rows.append(row_idx)

    if current_rows:
        sections.append(current_rows)

    return sections


def import_vouchers():
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active
    sections = parse_sections(ws)

    print(f"检测到 {len(sections)} 个月份段")
    # We expect 5 months (Jan-May). If more, use first 5.
    month_map = {0: 1, 1: 2, 2: 3, 3: 4, 4: 5}

    db = SessionLocal()

    # Verify company 1 exists
    company = db.query(Company).filter(Company.id == COMPANY_ID).first()
    if not company:
        print("❌ Company 1 not found!")
        return

    # Build account code lookup
    acct_codes = {}
    for a in db.query(Account).filter(Account.company_id == COMPANY_ID).all():
        acct_codes[a.code] = a

    # Delete old vouchers for company 1
    old_vouchers = db.query(Voucher).filter(Voucher.company_id == COMPANY_ID).all()
    for v in old_vouchers:
        db.query(VoucherEntry).filter(VoucherEntry.voucher_id == v.id).delete()
    db.query(Voucher).filter(Voucher.company_id == COMPANY_ID).delete()
    db.commit()
    print(f"已清除旧凭证数据")

    created = 0
    entries_created = 0
    unknown_codes = set()

    for sec_idx, rows in enumerate(sections):
        month = month_map.get(sec_idx, sec_idx + 1)
        if month > 5:
            break

        # Group rows by voucher number within this section
        voucher_rows = defaultdict(list)
        for row_idx in rows:
            vno = str(ws.cell(row=row_idx, column=2).value or "").strip()
            if vno:
                voucher_rows[vno].append(row_idx)

        print(f"\n{'=' * 60}")
        print(f"2026-{month:02d}月: {len(voucher_rows)} 张凭证, {len(rows)} 行分录")

        # Assign dates: distribute evenly across month
        day = 1
        for vno, vrows in sorted(voucher_rows.items()):
            # Parse voucher data
            entries_data = []
            voucher_total = 0

            for row_idx in vrows:
                summary = str(ws.cell(row=row_idx, column=6).value or "").strip()
                acct_code = str(ws.cell(row=row_idx, column=7).value or "").strip()
                debit = float(ws.cell(row=row_idx, column=10).value or 0)
                credit = float(ws.cell(row=row_idx, column=11).value or 0)
                aux = str(ws.cell(row=row_idx, column=9).value or "").strip()

                # Look up the exact account code
                code = acct_code
                if code not in acct_codes:
                    # Try to find by name if code not found
                    unknown_codes.add(code)
                    continue

                entries_data.append(
                    {
                        "account_code": code,
                        "debit": debit,
                        "credit": credit,
                        "description": summary,
                        "aux": aux,
                    }
                )
                voucher_total += debit  # both debit and credit should be same total

            if not entries_data:
                continue

            date = f"2026-{month:02d}-{day:02d}"
            day = min(day + 1, 28)  # increment day, cap at 28

            # Determine voucher type from analysis
            has_bank = any("1002" in e["account_code"][:4] for e in entries_data)
            voucher_type = "payment" if has_bank else "transfer"

            # Create voucher
            v = Voucher(
                company_id=COMPANY_ID,
                date=date,
                voucher_no=f"记字2026{month:02d}-{vno.split('-')[1].zfill(4)}",
                voucher_type=voucher_type,
                summary=entries_data[0]["description"],
                creator_id=1,
                status="draft",
            )
            db.add(v)
            db.flush()

            for e in entries_data:
                db.add(
                    VoucherEntry(
                        voucher_id=v.id,
                        account_code=e["account_code"],
                        debit=e["debit"],
                        credit=e["credit"],
                        description=e["description"],
                    )
                )
                entries_created += 1

            created += 1

    db.commit()

    print(f"\n{'=' * 60}")
    print(f"✅ 导入完成:")
    print(f"  凭证: {created} 张")
    print(f"  分录: {entries_created} 行")
    if unknown_codes:
        print(f"  ⚠️ 未知科目编码 ({len(unknown_codes)} 个):")
        for c in sorted(unknown_codes)[:20]:
            print(f"    - {c}")

    # Quick verification
    vouchers = db.query(Voucher).filter(Voucher.company_id == COMPANY_ID).all()
    total_d = sum(
        e.debit for v in vouchers for e in db.query(VoucherEntry).filter(VoucherEntry.voucher_id == v.id).all()
    )
    total_c = sum(
        e.credit for v in vouchers for e in db.query(VoucherEntry).filter(VoucherEntry.voucher_id == v.id).all()
    )
    print(f"\n  凭证合计: 借={total_d:,.2f} 贷={total_c:,.2f} {'✅' if abs(total_d - total_c) < 0.01 else '❌'}")

    db.close()


if __name__ == "__main__":
    import_vouchers()
